from openpyxl import load_workbook
import numpy as np

wb = load_workbook("./points_test.xlsx")

sheet = wb["Лист1"]
v_num = int(sheet['A1'].value)
print("Number of voters:", v_num)

p_num = int(sheet['A2'].value)
print("Number of participants:", p_num)

# voters list

voters = []

for i in range(2, v_num + 2):
    voters.append(str(sheet.cell(row=3, column=i).value))
print(voters)

# participants list

participants = []

for i in range (4, p_num + 4):
    participants.append(str(sheet.cell(row=i, column=1).value))
print(participants)

#

points_base = [[0] * p_num for i in range(v_num)]

for i in range(2, v_num + 2):
    for j in range (4, p_num + 4):
        temp = sheet.cell(row=j, column=i).value
        if temp is None:
            points_base[i - 2][j - 4] = 0
        else:
            points_base[i - 2][j - 4] = int(sheet.cell(row=j, column=i).value)

for i in range(0, v_num):
    print(points_base[i])

'''
current_points = [0] * p_num

for i in range(0, v_num):
    if voters[i] + " votes" in wb.sheetnames:
        wb.remove(wb[voters[i] + " votes"])
    new_sheet = wb.create_sheet(title=voters[i] + " votes")
    new_sheet['A1'] = voters[i]
    new_sheet['B1'] = "Given points"
    new_sheet['C1'] = "Sum of points"
    for j in range(0, p_num):
        new_sheet['A' + str(j + 2)] = participants[j]
        current_points[j] = current_points[j] + points_base[i][j]
        if points_base[i][j] != 0:
            new_sheet['B' + str(j + 2)] = points_base[i][j]
        new_sheet['C' + str(j + 2)] = current_points[j]
    new_sheet.auto_filter.ref = "A1:C" + str(p_num + 1)
    new_sheet.auto_filter.add_sort_condition("C1:C" + str(p_num + 1))

wb.save("./points_test.xlsx")

'''

current_points = [0] * p_num
temp_mtx = [[0] * p_num for i in range(3)]

for i in range(0, v_num):
    temp_arr = []
    if voters[i] + " votes" in wb.sheetnames:
        wb.remove(wb[voters[i] + " votes"])
    new_sheet = wb.create_sheet(title=voters[i] + " votes")
    new_sheet['A1'] = voters[i]
    new_sheet['B1'] = "Given points"
    new_sheet['C1'] = "Sum of points"
    new_sheet['E1'] = voters[i]
    new_sheet['F1'] = "Given points"
    new_sheet['G1'] = "Sum of points"
    for j in range(0, p_num):
        current_points[j] = current_points[j] + points_base[i][j]
        temp = (participants[j], points_base[i][j], current_points[j],)
        temp_arr.append(temp)
    data_type = [('part', 'U100'), ('get', int), ('sum', int)]
    sort_temp_arr = np.array(temp_arr, dtype=data_type)
    sort_temp_arr = np.sort(sort_temp_arr, order='sum') # sort
    res_arr = sort_temp_arr[::-1]

    print(voters[i], "voting:")
    print(res_arr)

    for j in range(0, p_num):
        if j < (p_num + 1) // 2:
            new_sheet['A' + str(j + 2)] = res_arr[j][0]
            if res_arr[j][1] != 0:
                new_sheet['B' + str(j + 2)] = res_arr[j][1]
            new_sheet['C' + str(j + 2)] = res_arr[j][2]
        else:
            new_sheet['E' + str(j % ((p_num + 1) // 2) + 2)] = res_arr[j][0]
            if res_arr[j][1] != 0:
                new_sheet['F' + str(j % ((p_num + 1) // 2) + 2)] = res_arr[j][1]
            new_sheet['G' + str(j % ((p_num + 1) // 2) + 2)] = res_arr[j][2]


    ''' вставление в таблицу
    for j in range(0, p_num):
        new_sheet['A' + str(j + 2)] = res_arr[j][0]
        if res_arr[j][1] != 0:
            new_sheet['B' + str(j + 2)] = res_arr[j][1]
        new_sheet['C' + str(j + 2)] = res_arr[j][2]

    new_sheet.auto_filter.ref = "A1:C" + str(p_num + 1)
    new_sheet.auto_filter.add_sort_condition("C1:C" + str(p_num + 1))
    '''

wb.save("./points_test.xlsx")
